#!/usr/bin/env python

import os, sys
import glob

"""
XML Spreadsheet format
"""
import openpyxl
from openpyxl import workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill, Alignment

"""
BIFF Spreadsheet format
"""
import xlrd

"""
Generate the consolidated codes
"""
import xlwt

"""
Extract content from an unknown state
"""

def si(content):
    """
    If it gets a None, return '', else return cleaned string
    """
    import types
    if isinstance(content, openpyxl.cell.Cell):
        if isinstance(content.value, types.NoneType):
            return unicode('')
        else:
            return unicode(content.value).strip()
    elif isinstance(content, xlrd.sheet.Cell):
        if content.ctype == xlrd.sheet.XL_CELL_EMPTY:
            return unicode('')
        else:
            return unicode(content.value).strip()
    return unicode(content).strip()

# General Dictionary of codes to columns
MAPPING_CODES = {'Mapping to BRIDG Defined Class' : 'BRIDG Defined Class C-Code',
                 'Mapping to BRIDG Defined Class Attribute' : 'BRIDG Defined Class Attribute C-Code',
                 'Mapping to BRIDG Performed Class' : 	'BRIDG Performed Class C-Code',
                 'Mapping to BRIDG Performed Class Attribute' : 'BRIDG Performed Class Attribute C-Code',
                 'Mapping to BRIDG Non-defined/Non-performed Class' : 'BRIDG Non-defined/Non-performed Class C-Code',
                 'Mapping to BRIDG Non-defined/Non-performed Class Attribute' : 'BRIDG Non-defined/Non-performed Class Attribute C-Code',
                 'Mapping to BRIDG Planned Class' : 'BRIDG Planned Class C-Code',
                 'Mapping to BRIDG Planned Class Attribute' : 'BRIDG Planned Class Attribute C-Code',
                 'Variable Name' : 'Variable Name C-Code',
                 'ISO 21090 Datatype' :	'ISO 21090 Datatype C-Code',
                 'ISO 21090 Datatype Constraint' : 'ISO 21090 Datatype Constraint C-Code',
                 'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES' : 'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES C-Code'}

MAPPING_ORDER = [u'Variable Name',
                 u'Variable Name C-Code',
                 u'Variable Label',
                 u'SHARE Generic Definition',
                 u'SDTM IG 3.1.2',
                 u'SEND 3.0',
                 u'CDASH V1.1',
                 u'CDASH V1.1 Conceptual Datatype',
                 u'SDTM IG 3.1.2 Datatype',
                 u'Codelist Master',
                 u'Set of Valid Values',
                 u'Assigned Value',
                 u'Mapping to BRIDG Defined Class',
                 u'Mapping to BRIDG Defined Class Attribute',
                 u'BRIDG Defined Class C-Code',
                 u'BRIDG Defined Class Attribute C-Code',
                 u'Mapping to BRIDG Performed Class',
                 u'Mapping to BRIDG Performed Class Attribute',
                 u'BRIDG Performed Class C-Code',
                 u'BRIDG Performed Class Attribute C-Code',
                 u'Mapping to BRIDG Non-defined/Non-performed Class',
                 u'Mapping to BRIDG Non-defined/Non-performed Class Attribute',
                 u'BRIDG Non-defined/Non-performed Class C-Code',
                 u'BRIDG Non-defined/Non-performed Class Attribute C-Code',
                 u'Mapping to BRIDG Planned Class',
                 u'Mapping to BRIDG Planned Class Attribute',
                 u'BRIDG Planned Class C-Code',
                 u'BRIDG Planned Class Attribute C-Code',
                 u'ISO 21090 Datatype',
                 u'ISO 21090 Datatype C-Code',
                 u'ISO 21090 Datatype Component',
                 u'AsCollectedIndicator',
                 u'Observation, ObservationResult, Activity, Relationship',
                 u'Description of Observation, ObservationResult or Activity or Relationship - CODED VALUES',
                 u'Description of Observation, ObservationResult or Activity or Relationship - C-Codes',
                 u'Description of Observation, ObservationResult or Activity or Relationship - NON-CODED VALUES',
                 u'NOTES',
                 u'Null flavors',
                 u'Boolean Mapping',
                 u'ISO 21090 Datatype Constraint',
                 u'ISO 21090 Datatype Constraint C-Code',
                 u'ISO 21090 Datatype Constraint Attribute',
                 u'Observation or Activity'
                 ]

def vsort(this, that):
    """
    Variable name sort
    """
    if this.startswith('--') and that.startswith('--'):
        return cmp(this[2:].lower(), that[2:].lower())
    elif this.startswith('--'):
        return cmp(this[1:].lower(), that.lower())
    elif that.startswith('--'):
        return cmp(this.lower(), that[1:].lower())
    else:
        return cmp(this.lower(), that.lower())

class CodeLoader(object):
    
    def __init__(self, options):
        if options.template:
            self.terminology = self._load_template(options.template)
        else:
            self.terminology = self._load_reference(options.reference)
        
    def _load_template(self, filename):
        """
        Template has terminology in the 'Terms' sheet
        """
        terminology = {}
        print 'Loading %s from Content Template' % filename
        try:
            workbook = xlrd.open_workbook(filename)
        except xlrd.biffh.XLRDError, e:
            # use the xml reader
            terminology = self._load_template_xml(filename)
            return terminology
        except IOError, e:
            print 'Failed to open %s - %s' % (filename, e)
            sys.exit(1)
        if 'Fields to be Coded' in workbook.sheet_names():
            terminology = self._load_unique_items_to_code(workbook)
        else:
            sheet = workbook.sheet_by_name('Terms')
            for row in range(1, sheet.nrows):
                content = sheet.row_values(row)
                terminology[si(content[0])] = si(content[1])
        print 'Loaded %s terms' % len(terminology)
        return terminology

    def _load_template_xml(self, filename):
        """
        use a different reader
        """
        try:
            workbook = openpyxl.reader.excel.load_workbook(filename)
        except Exception, e:
            import traceback, sys
            print 'Failed to open %s : %s' % (filename, e)
            traceback.print_tb(sys.exc_info()[2])
            return
        terminology = {}
        if 'Fields to be Coded' in workbook.get_sheet_names():
            terminology = self._load_unique_items_to_code(workbook)
        else:    
            sheet = workbook.get_sheet_by_name('Terms')
            for (idx, row) in enumerate(sheet.rows):
                if idx == 0:
                    # ignore header
                    continue
                if not si(row[0]):
                    # blank rows
                    continue
                terminology[si(row[0])] = {None : ''}.get(si(row[1]), si(row[1]))
        print 'Loaded %s terms' % len(terminology)
        return terminology

    def _load_unique_items_to_code(self, workbook):
        """
        Load in the terms
        TODO: Refactor this!!
        """
        terminology = {}
        if isinstance(workbook, openpyxl.workbook.Workbook):
            for sheet_name in ('Fields to be Coded', 'Fields Coded'):
                sheet = workbook.get_sheet_by_name(sheet_name)
                for (idx, content) in enumerate(sheet.rows):
                    if idx == 0:
                        # ignore header
                        continue
                    if not si(content[0]):
                        # blank rows
                        continue
                    if '|' in si(content[0]):
                        # multiple terms, split 'em
                        terms = [x.strip() for x in si(content[0]).split('|')]
                        values = [x.strip() for x in si(content[1]).split('|')]
                        if not len(terms) == len(values):
                            print 'Error in mapping: %s => %s (Not a match in numbers)' % (si(content[0]), si(content[1]))
                        else:
                            terminology.update(dict(zip(terms, values)))
                        pass
                    else:
                        # Map a 'None' to ''
                        terminology[si(content[0])] = {None : '',
                                                       'None' : ''}.get(si(content[1]), si(content[1]))
            
        else:
            for sheet_name in ('Fields to be Coded', 'Fields Coded'):
                sheet = workbook.sheet_by_name(sheet_name)
                for row in range(1, sheet.nrows):
                    content = sheet.row_values(row)
                    if '|' in si(content[0]):
                        terms = [x.strip() for x in si(content[0]).split('|')]
                        values = [x.strip() for x in si(content[1]).split('|')]
                        if not len(terms) == len(values):
                            print 'Error in mapping: %s => %s (Not a match in numbers)' % (si(content[0]), si(content[1]))
                        else:
                            terminology.update(dict(zip(terms, values)))
                    else:
                        terminology[si(content[0])] = {None : '', 'None' : ''}.get(si(content[1]), si(content[1]))
        return terminology
        
    def _load_reference(self, filename):
        terminology = {}
        # Office 2003 or later xml format 
        print 'Loading %s as a reference' % filename
        try:
            workbook = openpyxl.reader.excel.load_workbook(filename)
        except Exception, e:
            print 'Failed to open %s : %s (Strike 1)' % (filename, e)
            # use the xlrd module
            terminology = self._load_reference_xls(filename)
            return terminology
        for sheet_name in workbook.get_sheet_names():
            
            if (not 'Generic' in sheet_name):
                # Only look for headings in pages we don't care about
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            for (idx, row) in enumerate(sheet.rows):
                if not row[0].value:
                    # blank rows
                    continue
                
                if row[0].value.strip().lower() == 'variable name':

                    # Found the row with the column headings                                                         
                    for this_row in sheet.rows[idx + 1:]:
                        if this_row[0].value == '' or this_row[0].value is None:
                            continue
                        if this_row[1].value is None:
                            print 'C-code required for %s' % this_row[0].value.strip()
                            terminology[str(this_row[0].value).strip()] = ''
                        else:
                            if str(this_row[1].value).strip() in terminology.items():
                                print 'Duplicated C-code: %s' % str(this_row[1].value).strip()
                            terminology[str(this_row[0].value).strip()] = str(this_row[1].value).strip()
                    else:
                        return terminology

    def _load_reference_xls(self, filename):
        terminology = {}
        # Office 2003 or later xml format 
        print 'Loading %s as a reference' % filename
        try:
            workbook = xlrd.open_workbook(filename)
        except Exception, e:
            print 'Failed to open %s : %s (Last chance!!)' % (filename, e)
            sys.exit(1)
        for sheet_name in workbook.get_sheet_names():
            
            if (not 'Generic' in sheet_name):
                # Only look for headings in pages we don't care about
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            for (idx, row) in enumerate(sheet.rows):
                if not row[0].value:
                    # blank rows
                    continue
                
                if row[0].value.strip().lower() == 'variable name':

                    # Found the row with the column headings                                                         
                    for this_row in sheet.rows[idx + 1:]:
                        if this_row[0].value == '' or this_row[0].value is None:
                            continue
                        if this_row[1].value is None:
                            print 'C-code required for %s' % this_row[0].value.strip()
                            terminology[str(this_row[0].value).strip()] = ''
                        else:
                            if str(this_row[1].value).strip() in terminology.items():
                                print 'Duplicated C-code: %s' % str(this_row[1].value).strip()
                            terminology[str(this_row[0].value).strip()] = str(this_row[1].value).strip()
                    else:
                        return terminology
        

    def column_sort(self, cola, colb):
        return cmp(MAPPING_ORDER.index(cola), MAPPING_ORDER.index(colb))
                   
    def dump_map(self, filename):
        coded_columns = MAPPING_CODES.keys()
        boldtype = xlwt.easyxf("font: bold on; align: wrap on, vert top, horiz center; borders: left 1, top 1, bottom 1, right 1;")
        borderedtype = xlwt.easyxf("borders: left 1, top 1, bottom 1, right 1; align: vert top, horiz left;")
        try:
            coded = xlwt.Workbook()
        except xlwt.Exception:
            print "Can't output"
            sys.exit()
            # Office 2003 or later xml format
        try:
            workbook = openpyxl.load_workbook(filename)
        except Exception, e:
            import traceback
            print 'Failed to open %s : %s' % (filename, e)
            traceback.print_tb(sys.exc_info()[2])
            return
        # flag for refresh
        DIRTY = []
        for sheet_name in workbook.get_sheet_names():
            if ('Rules' in sheet_name or 
                'Decision' in sheet_name or 
                'Terminology' in sheet_name or
                'README' in sheet_name or
                'Null' in sheet_name):
                # Only look for headings in pages we don't care about
                continue
            coded_tab = coded.add_sheet(sheet_name)
            sheet = workbook.get_sheet_by_name(sheet_name)
            for (offset, row) in enumerate(sheet.rows, 1):
                if not si(row[0]):
                    # blank rows
                    continue
                # look where to start coding
                if si(row[0]).lower() == 'variable name':
                    # Found the row with the column headings
                    headers = [si(x) for x in row]
                    # filter the coded_columns on columns where the coded column exists in the dataset
                    filtered_columns = filter(lambda x: MAPPING_CODES.get(x) in headers, coded_columns)
                    # Define the mapped headers
                    overlap = sorted(set(headers).intersection(set(filtered_columns)), cmp=self.column_sort)
                    # Write the headers for the coded columns
                    for (idx, col) in enumerate(overlap, 1):
                        coded_tab.write(0, idx, MAPPING_CODES.get(col), boldtype)
                        # Iterate from the variable_name row forward
                    for (offset_row, this_row) in enumerate(sheet.rows[offset:], 1):
                        # Put the content into a dictionary
                        mapped_content = dict(zip(headers, [si(x) for x in this_row]))
                        # if the first column is blank, skip
                        if mapped_content.get('Variable Name') in ['', None]:
                            continue
                        # Write the Variable as a Pivot
                        coded_tab.write(offset_row, 0, mapped_content.get('Variable Name'), borderedtype)
                        for (idx, col) in enumerate(overlap, 1):
                            if mapped_content.get(col) not in [None, '', 'na']:
                                if '|' in mapped_content.get(col):
                                    # Multi-value columns
                                    terms = [x.strip() for x in mapped_content.get(col).split('|')]
                                    codes = [self.terminology.get(x, '') for x in terms]
                                    proposed = ' | '.join(codes)
                                else:
                                    proposed = self.terminology.get(mapped_content.get(col, ""), '')
                                # check and see if the code to be assigned, is already assigned so we don't have to do so
                                # much work ;-)
                                if mapped_content.get(MAPPING_CODES.get(col)) != proposed:
                                    # TODO: update the code in the scanned sheet here!
                                    target = sheet.cell("%s%s" % (get_column_letter(headers.index(MAPPING_CODES.get(col))), offset_row + offset))
                                    target.style.alignment.vertical = Alignment.VERTICAL_TOP
                                    target.value = proposed
                                    DIRTY.append("Sheet: %s; Variable: %s; Column %s was updated" % (sheet_name, mapped_content.get('Variable Name'), col))
                                coded_tab.write(offset_row, idx, proposed, borderedtype)
                            else:
                                coded_tab.write(offset_row, idx, '', borderedtype)
        if len(DIRTY) != 0:
            print "Content to be updated:"
            print "\n".join(DIRTY)
            workbook.save("%s_WITH_CODES.xls" % (os.path.splitext(filename)[0]))
            coded.save("%s_CODED.xls" % (os.path.splitext(filename)[0]))
            
def template_sort(x, y):
    """
    Custom sort for Content Templates
    """
    try:
      (dx, tx) = x.split()[:2]
      (dy, ty) = y.split()[:2]
    except ValueError:
      print "Error on template sort for: %s %s" % (x, y)
      sys.exit()
    # Generic for both
    if len(tx) > 2 and len(ty) > 2:
        return cmp(tx, ty)
    # x is Generic
    elif len(tx) > 2:
        return 1
    # y is Generic
    elif len(ty) > 2:
        return -1
    # both non-Generic
    else:
        return cmp(tx, ty)

OBS_CLASS = {"INTERVENTIONS" : ["EX", "CM", "SU"],
              "FINDINGS" : ["EG", "IE", "LB", "PE", "QS", "SC",
               "VS", "MS", "MB", "DA", "PC", "PP"],
              "EVENTS" : ["AE", "DS", "MH", "DV", "CE"],
              "SPECIALPURPOSE" : ["DM", "SV", "SE", "CO"]}
              
class CodeExtractor(object):
    """
    Extract Items that need to be coded out
    """
    def __init__(self):
        self.coded_entries = {}

    def load_files(self, xlfiles=[]):
        """
        Load in the files
        """
        if xlfiles == []:
            import glob
            xlfiles = glob.glob("*Template.xls*")
        for xlfile in xlfiles:
            # no temp files
            if xlfile.startswith('~'):
                continue
            # fence btw biff-based and XML-based
            if os.path.splitext(xlfile)[1] == ".xlsx":
                self._load_template_xml(xlfile)
            else:
                self._load_template(xlfile)

        self._export_codes()

    def _extract_sheets(self):
        """
        Pull out the sheets in which an item has been referenced
        """
        holders = {}
        for item in self.coded_entries.itervalues():
            for holder in item.holders:
                holders.setdefault(holder, '')
        return sorted(holders.keys(), cmp=template_sort)

    def _export_codes(self):
        """
        Report the findings in an excel spreadsheet
        """
        boldtype = xlwt.easyxf("font: bold on; align: wrap on, vert centre, horiz center")
        book = xlwt.Workbook()
        """
        Fields that have no code assigned anywhere
        """
        tocode = book.add_sheet('Fields to be Coded')

        all_items = sorted([x for x in self.coded_entries.itervalues()])
        tocode.write(0, 0, "Field", boldtype)
        tocode.write(0, 1, "Code", boldtype)
        idx = 1
        for item in all_items:
            if item.uncoded or item.new:
                tocode.write(idx, 0, item.name)
                idx = idx + 1
        """
        Fields that have no code assigned anywhere
        """
        toresolve = book.add_sheet('Fields to be Resolved')
        toresolve.write(0, 0, "Field", boldtype)
        toresolve.write(0, 1, "Codes", boldtype)
        for (idx, item) in enumerate(all_items, 1):
            if item.conflicts:
                print "%s is conflicted: %s" % (item.name, item.coded)
                from operator import itemgetter
                # Item Name
                toresolve.write(idx, 0, item.name)
                _codes = sorted(item.coded, key=itemgetter(1))
                for (c_ind, (cf, va)) in enumerate(_codes):
                    toresolve.write(idx, (c_ind*2)+1, cf)
                    toresolve.write(idx, (c_ind*2)+2, va)

        """
        Fields that have no code assigned anywhere
        """
        nocode = book.add_sheet('Fields Coded')
        nocode.write(0, 0, "Field", boldtype)
        nocode.write(0, 1, "Code", boldtype)
        coded_state = [x for x in all_items if not (x.uncoded or x.conflicts)]
        for (idx, item) in enumerate(coded_state, 1):
            if item.uncoded or item.conflicts:
                continue
            nocode.write(idx, 0, item.name)
            nocode.write(idx, 1, item.coded)
        """
        Instances
        """
        instances = book.add_sheet('Instances')
        instances.write(0, 0, "Field", boldtype)
        instances.write(0, 1, "Context", boldtype)
        all_sections = self._extract_sheets()
        # Write out the file names
        for (idx, holder) in enumerate(all_sections):
            instances.write(0, 2 + idx, holder, boldtype)  
        for (idx, item) in enumerate(sorted(all_items), 1):
            instances.write(idx, 0, item.name)
            instances.write(idx, 1, item.context)
            for film in item.holders:
                jdx = 2 + all_sections.index(film)
                instances.write(idx, jdx, 'X')
        for (obsclass, domain) in OBS_CLASS.iteritems():
          pass
        import time
        book.save('SHARE_Unique_Items_To_Code_%s.xls' % time.strftime('%Y%m%d'))
            
        
    def _extract_code_couples(self, headers):
        """
        Extract couples from the headers
        """
        couples = {}
        import copy
        _dup = copy.deepcopy(headers)
        for column in headers:
            if 'c-code' in column.lower():
                # location of c-code
                r_indx = column.lower().index('c-code')
                prefix = ''.join(column[:(r_indx-1)])
                for _t in _dup:
                    if prefix.strip().lower() in _t.strip().lower():
                        if (column.strip().lower() != _t.strip().lower()):
                            couples[_t] = column
                        else:
                            continue
                    if _t == column:
                        print "No source column found for %s" % column
                        break
        print "Column maps"
        for (code, column) in couples.iteritems():
            print "%s -> %s" % (code, column)
        return couples
    
    def extract_code_couples(self, headers):
        return MAPPING_CODES

    def _load_template(self, filename):
        """
        Template has terminology in the 'Terms' sheet
        """
        print "Parsing %s" % filename
        try:
            workbook = xlrd.open_workbook(filename)
        except xlrd.biffh.XLRDError, e:
            print 'Failed to open %s - %s' % (filename, e)
            sys.exit(1)
        for sheet_name in workbook.sheet_names():
            if ('Rules' in sheet_name or
                'Decision' in sheet_name or
                'Terminology' in sheet_name or
                'Flavors' in sheet_name):
                # Don't look for headings in pages that we don't care about
                continue
            sheet = workbook.sheet_by_name(sheet_name)
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                if not si(row[0]):
                    # blank rows
                    continue
                if si(row[0]).lower() == 'variable name':
                    headers = [si(x) for x in row]
                    couples = self.extract_code_couples(headers)
                    # Found the row with the column headings
                    for this_row_idx in range(row_idx + 1, sheet.nrows):
                        this_row = sheet.row(this_row_idx)
                        if si(this_row[0]) == '':
                            continue
                        _content = dict(zip(headers, [si(x) for x in this_row]))
                        _name = _content.get('Variable Name')
                        if not _name:
                            continue
                        for (c_key, c_code) in couples.iteritems():
                            sourcecol = _content.get(c_key)
                            codecol = _content.get(c_code)
                            if not sourcecol in [None, "", "na"]:
                                item = self.coded_entries.setdefault(sourcecol, CodedEntry(sourcecol))
                                item.add_code(c_key, codecol, filename)
        
    def _load_template_xml(self, filename):
        """
        use a different reader (Later versions of XL using the XML format)
        """
        print "Parsing %s" % filename
        try:
            workbook = openpyxl.reader.excel.load_workbook(filename)
        except Exception, e:
            import traceback, sys
            print 'Failed to open %s : %s' % (filename, e)
            traceback.print_tb(sys.exc_info()[2])
            return
        for sheet_name in workbook.get_sheet_names():
            if ('Rules' in sheet_name or
                'Decision' in sheet_name or
                'Terminology' in sheet_name or
                'Flavors' in sheet_name):
                # Don't look for headings in pages that we don't care about
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            for (idx, row) in enumerate(sheet.rows):
                if not row[0].value:
                    # blank rows
                    continue
                if si(row[0]).lower() == 'variable name':
                    headers = [si(x) for x in row]
                    couples = self.extract_code_couples(headers)
                    # print 'Headers: %s' % headers
                    # Found the row with the column headings
                    for this_row in sheet.rows[idx + 1:]:
                        if si(this_row[0]) == '' or si(this_row[0]) is None:
                            continue
                        _content = dict(zip(headers, [si(x) for x in this_row]))
                        _name = _content.get('Variable Name')
                        if not _name:
                            continue
                        for (c_key, c_code) in couples.iteritems():
                            try:
                                sourcecols = [x.strip() for x in _content.get(c_key).split('|')]
                                codecols = [x.strip() for x in _content.get(c_code).split('|')]
                            except AttributeError:
                                continue
                            for (c_indx, sourcecol) in enumerate(sourcecols):
                                if not sourcecol in [None, "", "na"]:
                                    item = self.coded_entries.setdefault(sourcecol, CodedEntry(sourcecol))
                                    try:
                                        item.add_code(c_key, codecols[c_indx], filename)
                                    except IndexError:
                                        item.add_code(c_key, "", filename)


class CodedEntry(object):
    """
    Holder for something requiring coding
    """
    def __init__(self, name):
        self.name = name
        """
        Only set if there is a single code
        """
        self.code = {}
        self.holders = []

    def add_code(self, context, code, holder):
        """
        Context = Column Heading 
        Code = Any code assigned - can be None
        nasty, yucky branchy code... bleurgh
        """
        if not holder in self.holders:
            self.holders.append(holder)
        # Map a None to a missing string - easier to deal with
        _code = {None : ''}.get(code, code)
        # not currently defined or equal to ""
        if not self.code.get(context):
            # Not currently defined - set it
            self.code[context] = _code
            return
        # Currently defined (maybe blank)
        current = self.code.get(context)
        # If currently blank
        if (current == ""):
            # Missing is replaced by non-missing
            if _code != "":
                self.code[context] = _code
                return
        else:
            if _code == "":
                return
            if current != _code:
                print "Redefinition of %s" % self.name
                    # do something
                spl = current.split(',')
                if not _code in spl:
                    spl.append(_code)
                    self.code[context] = ','.join(spl)
                

    @property
    def context(self):
        """
        Looks at the source Column's and derive the context (VAR|BRIDG|ISO)
        """
        response = []
        for column in self.code.keys():
            if column.startswith('ISO'):
                if not 'ISO' in response:
                    response.append('ISO')
            elif column.startswith('Mapping'):
                if not 'BRIDG' in response:
                    response.append('BRIDG')
            elif column.startswith('Variable'):
                if not 'VAR' in response:
                    response.append('VAR')
        return ','.join(response)

    @property
    def uncoded(self):
        """
        True - only value is None
        False - just the one
        """
        for i in self.code.itervalues():
            if i.strip() != '':
                return False
        return True

    @property
    def conflicts(self):
        """
        True - more than one code for self.name
        False - just the one
        """
        _values = []
        for i in self.code.itervalues():
            if not i in _values:
                if i != '':
                    _values.append(i)
        if len(_values) > 1:
            return True
        return False

    @property
    def new(self):
        return "CNEW" in self.code.values()
            
    @property
    def coded(self):
        """
        Single code
        """
        if self.uncoded:
            return ''
        elif self.conflicts:
            """
            returns a list
            """
            return [x for x in self.code.iteritems()]
        else:
            for i in self.code.itervalues():
                if i != "":
                    return i

    def __cmp__(self, other):
        if self.name.startswith('--') and other.name.startswith('--'):
            return cmp(self.name[2:].lower(), other.name[2:].lower())
        elif self.name.startswith('--'):
            return cmp(self.name[2:].lower(), other.name.lower())
        elif other.name.startswith('--'):
            return cmp(self.name.lower(), other.name[2:].lower())
        else:
            return cmp(self.name.lower(), other.name.lower())

if __name__ == "__main__":

    import optparse
    parser = optparse.OptionParser()
    parser.add_option('-r',
                      help="Reference File",
                      metavar='FILE',
                      action='store',
                      dest='reference',
                      default='')
    parser.add_option('-t',
                      help="Template File (Populated with Terminology)",
                      metavar='FILE',
                      action='store',
                      dest='template',
                      default='')
    parser.add_option('-x',
                      action='store_true',
                      default=False,
                      dest='extract',
                      help="Extract codes that require coding")
    (opts, args) = parser.parse_args()
    if opts.extract:
        extractor = CodeExtractor()
        extractor.load_files(args)
        sys.exit()
    if not ((opts.reference != '') ^ (opts.template != '')):
        sys.exit()
    tk = CodeLoader(opts)
    for template in glob.glob("*Template.xlsx"):
        if template.startswith("~"):
            continue
        print 'Generating %s' % template
        tk.dump_map(template)

