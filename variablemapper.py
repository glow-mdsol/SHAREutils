#!/usr/bin/env python

import os
import sys
import openpyxl
import xlrd
import xlwt
import time

class ContentReader(object):

    def __init__(self):
        self.files = []
        self.variable_names = {}

    def load_spreadsheet(self, filename):
        if os.path.splitext(filename)[1] not in ['.xlsx', '.xls']:
            print 'Ignoring %s' % filename
            return
        print 'Opening %s' % filename
        self.files.append(filename)
        if os.path.splitext(filename)[1] == '.xlsx':
            self._xlsx_loader(filename)
        else:
            self._xls_loader(filename)
        print 'Loaded %s variables' % len(self.variable_names)

    def export_content(self):
        boldtype = xlwt.easyxf("font: bold on; align: wrap on, vert centre, horiz center")
        book = xlwt.Workbook()
        info = book.add_sheet('Information')
        info.write(0, 0, 'Variable Name', boldtype)
        for (f_idx, filename) in enumerate(sorted(self.files)):
            info.write(0, f_idx + 1, filename, boldtype)
        for (idx, term) in enumerate(sorted(self.variable_names.values())):
            info.write(idx + 1, 0, term.name)
            for filename in sorted(term.files):
                info.write(idx + 1, 1 + sorted(self.files).index(filename), 'X', boldtype)

        sheet = book.add_sheet('Terms')
        sheet.write(0, 0, 'Variable Name', boldtype)
        sheet.write(0, 1, 'Code', boldtype)
        #sheet.write(0, 2, 'Definition', boldtype)
        for (idx, term) in enumerate(sorted(self.variable_names.values())):
            sheet.write(idx + 1, 0, term.name)
            sheet.write(idx + 1, 1, term.code)
            #sheet.write(idx + 1, 2, term.definition)

        sheet_code = book.add_sheet('Codes')
        sheet_code.write(0, 0, 'Variable Name', boldtype)
        for (f_idx, filename) in enumerate(sorted(self.files)):
            sheet_code.write(0, f_idx + 1, filename, boldtype)
        for (idx, term) in enumerate(sorted(self.variable_names.values())):
            sheet_code.write(idx + 1, 0, term.name)
            for (filename, code) in term.codes.items():
                offset = sorted(self.files).index(filename)
                sheet_code.write(idx + 1, offset + 1, code)

        # sheet_def = book.add_sheet('Definitions')
        # sheet_def.write(0, 0, 'Variable Name', boldtype)
        # for (f_idx, filename) in enumerate(sorted(self.files)):
        #     sheet_def.write(0, f_idx + 1, filename, boldtype)
        # for (idx, term) in enumerate(sorted(self.variable_names.values())):
        #     sheet_def.write(idx + 1, 0, term.name)
        #     for (filename, definition) in term.definitions.items():
        #         offset = sorted(self.files).index(filename)
        #         sheet_def.write(idx + 1, offset + 1, definition)

        book.save('SHARE_Combined_Variables_%s.xls' % time.strftime('%Y%m%d'))

    def variable_report(self):
        boldtype = xlwt.easyxf("font: bold on; align: wrap on, vert centre, horiz center")
        wrappable = xlwt.easyxf("align: wrap on, vert top, horiz left")
        book = xlwt.Workbook()
        # create the location sheet (where in all the files I've opened is the fragment)
        info = book.add_sheet('Variable Locations')
        info.write(0, 0, "Variable", boldtype)
        _sf = sorted(self.files)
        for (colidx, filename) in enumerate(_sf, 1):
            info.write(0, colidx, filename, boldtype)
        _sv = sorted(self.variable_names.values())
        for (rowidx, variable) in enumerate(_sv, 1):
            info.write(rowidx, 0, variable.name)
            for filename in variable.filenames:
                info.write(rowidx, _sf.index(filename) + 1, "X")
        # next the definitions
        defs = book.add_sheet('Variable Definitions')
        for idx in range(1, 9):
            if idx % 2 == 0:
                defs.write(0, idx, 'Locations', boldtype)
            else:
                defs.write(0, idx, 'Generic Definition', boldtype)
            defs.col(idx).width = 256 * 35
        defs.write(0, 0, "Variable Name", boldtype)
        for (rowidx, variable) in enumerate(_sv, 1):
            defs.write(rowidx, 0, variable.name)
            for (colidx, (definition, locations)) in enumerate(variable.definitions.iteritems(), 1):
                defs.write(rowidx, ((2 * colidx) - 1), definition, wrappable)
                defs.write(rowidx, ((2 * colidx)), ';'.join(sorted([os.path.splitext(x)[0] for x in locations])), wrappable)
        
        # then the labels
        labs = book.add_sheet('Variable Labels')
        for idx in range(1, 7):
            labs.col(idx).width = 256 * 35
            if idx % 2 == 0:
                labs.write(0, idx, 'Locations', boldtype)
            else:
                labs.write(0, idx, 'Label', boldtype)
        labs.write(0, 0, "Variable Name", boldtype)
        for (rowidx, variable) in enumerate(_sv, 1):
            labs.write(rowidx, 0, variable.name)
            for (colidx, (label, locations)) in enumerate(variable.labels.iteritems(), 1):
                labs.write(rowidx, ((2 * colidx) - 1), label, wrappable)
                labs.write(rowidx, ((2 * colidx)), ';'.join(sorted([os.path.splitext(x)[0] for x in locations])), wrappable)
        book.save('SHARE_Variable_Report_%s.xls' % time.strftime('%Y%m%d'))
        
    def _xlsx_loader(self, filename):
        """
        Using pyopenxl
        """
        try:
            workbook = openpyxl.reader.excel.load_workbook(filename)
        except Exception, e:
            import traceback
            print 'Failed to open %s : %s' % (filename, e)
            traceback.print_tb(sys.exc_info()[2])
            return
        for sheet_name in workbook.get_sheet_names():
            if ('Rules' in sheet_name or
                'Decision' in sheet_name or
                'Terminology' in sheet_name or
                'Flavors' in sheet_name):
                # Don't look for headings in pages that we don't care about
                continue
            sheet = workbook.get_sheet_by_name(sheet_name)
            for (idx, row) in enumerate(sheet.rows):
                if not row[0].value:
                    # blank rows
                    continue
                if si(row[0]).lower() == 'variable name':
                    headers = [si(x) for x in row]
                    #print 'Headers: %s' % headers
                    # Found the row with the column headings
                    for this_row in sheet.rows[idx + 1:]:
                        if this_row[0] == '' or this_row[0] is None:
                            continue
                        _content = dict(zip(headers, [si(x) for x in this_row]))
                        _name = _content.get('Variable Name')
                        if not _name:
                            continue
                        vble = self.variable_names.setdefault(_name, VariableName(_name))
                        vble.load_row(_content, filename)

    def _xls_loader(self, filename):
        """
        using xlrd
        """
        # old school xls file
        try:
            workbook = xlrd.open_workbook(filename)
        except Exception, e:
            print 'Failed to open %s; %s' % (filename, e)
            return
        for sheet_name in workbook.sheet_names():
            if ('Rules' in sheet_name or
                'Decision' in sheet_name or
                'Terminology' in sheet_name or
                'Flavors' in sheet_name):
                # Don't look for headings in pages that we don't care about
                continue
            sheet = workbook.sheet_by_name(sheet_name)
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                if not si(row[0]):
                    # blank rows
                    continue
                if si(row[0]).lower() == 'variable name':
                    headers = [si(x) for x in row]
                    # Found the row with the column headings
                    for this_row_idx in range(row_idx + 1, sheet.nrows):
                        this_row = sheet.row(this_row_idx)
                        if si(this_row[0]) == '':
                            continue
                        _content = dict(zip(headers, [si(x) for x in this_row]))
                        _name = _content.get('Variable Name')
                        if not _name:
                            continue
                        vble = self.variable_names.setdefault(_name, VariableName(_name))
                        vble.load_row(_content, filename)


def si(content):
    """
    If it gets a None, return '', else return cleaned string
    """
    import types
    if isinstance(content, openpyxl.cell.Cell):
        if isinstance(content.value, types.NoneType):
            return unicode('')
        else:
            return unicode(content.value).strip()
    elif isinstance(content, xlrd.sheet.Cell):
        if content.ctype == xlrd.sheet.XL_CELL_EMPTY:
            return unicode('')
        else:
            return unicode(content.value).strip()
    return unicode(content).strip()


class VariableName(object):
    
    def __init__(self, name):
        # pass in the row as a dictionary agin the headers
        self.name = name
        self._definitions = {}
        self._labels = {}
        self._codes = {}

    def _collate_items(self, items):
        """
        group items that are the same together
        """
        collated = {}
        for (filename, item) in items.iteritems():
            collated.setdefault(item, []).append(filename)
        return collated

    @property
    def filenames(self):
        return set(self._labels.keys() + self._definitions.keys() + self._codes.keys())
    
    @property
    def labels(self):
        return self._collate_items(self._labels)

    @property
    def definitions(self):
        return self._collate_items(self._definitions)
                                                 
    @property
    def code(self):
        return ','.join(dict().fromkeys(self._codes.values()).keys())

    def __cmp__(self, other):
        if self.name.startswith('--') or other.name.startswith('--'):
            # should order --XXXX, then VSXXXX, etc
            if self.name[2:] == other.name[2:]:
                # same fragment
                if self.name.startswith('--'):
                    return -1
                else:
                    return 1
        return cmp(self.name, other.name)

    @property
    def locations(self):
        return set(self._labels.keys() + self._definitions.keys() + self._codes.keys())
    
    def load_row(self, row, filename):
        # pattern match on the keys for the definitions and c-codes (variable names)
        keys = row.keys()
        code = self._has_codes(keys)
        if code:
            _code = row.get(code)
            self._codes[filename] = _code
        definition = self._has_definition(keys)
        if definition != '':
            self._definitions[filename] = row.get(definition)
        label = self._has_label(keys)
        if label != '':
            self._labels[filename] = row.get(label)

    def _has_codes(self, keys):
        """
        Keys are the dict keys
        """
        for key in keys:
            if (key.strip().lower().startswith('variable name') and 'c-code' in key.strip().lower()):
                return key
        return ''

    def _has_definition(self, keys):
        """
        Keys are the dict keys
        """
        for key in keys:
            if 'definition' in key.lower():
                return key
        return ''

    def _has_label(self, keys):
        """
        Keys are the dict keys
        """
        for key in keys:
            if 'label' in key.lower():
                return key
        return ''

if __name__ == "__main__":
    import glob
    vmap = ContentReader()
    for excel in glob.glob("201*.xls*"):
        vmap.load_spreadsheet(excel)
        #vmap.export_content()
    vmap.variable_report()
